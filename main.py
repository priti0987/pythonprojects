import openpyxl
import easygui

data_sheet_path = "C:\\KhanBankApp\\TestData\\DataSheet.xlsx"
workbook = openpyxl.load_workbook(data_sheet_path)
sheet = workbook.get_sheet_by_name("Datasheet_Signup")
rows=(sheet.max_row)
col = (sheet.max_column)

for r in range(3, rows + 1):
    R = sheet.cell(r,1)
    E = sheet.cell(r,2)
    P = sheet.cell(r,3)
    print(R.value)
    print(E.value)
    print(P.value)
